import openpyxl
import random

namelist=['Дубровина Варвара','Рябова Ульяна','Карташов Давид','Кузнецов Николай','Белоусов Григорий','Максимов Егор','Иванов Михаил','Аникина Ева','Афанасьев Павел','Кузнецова Елизавета']

#заполняем столбец ФИО
def FIO ():
    for i in range(len(namelist)):
        value = namelist[i]
        cell = sheet.cell(row = i+2, column = 1)
        cell.value = value
    
# создаем новый excel-файл
wb = openpyxl.Workbook()

# добавляем новый лист
wb.create_sheet(title = 'Взносы', index = 0)

# получаем лист, с которым будем работать
sheet = wb['Взносы']

#Делаем строчечки больше ото фамилии не влезают,жирные они слишком
sheet.column_dimensions['A'].width = 25

#Вводим названия колонок
value = "ФИО"
cell = sheet.cell(row = 1, column = 1)
cell.value = value

value = "Взнос"
cell = sheet.cell(row = 1, column = 2)
cell.value = value

value = "Сумма"
cell = sheet.cell(row = 1, column = 3)
cell.value = value

value = "Месяц"
cell = sheet.cell(row = 1, column = 4)
cell.value = value

FIO ()



#Заполнение столбцов Взносы,Сумма
for i in range(len(namelist)):
    a=random.randint(0,1)
    if (a==1):
        value = "+"
        cell = sheet.cell(row = i+2, column =2)
        cell.value = value
        value = "35"
        cell = sheet.cell(row = i+2, column =3)
        cell.value = value
    else:
        value = "-"
        cell = sheet.cell(row = i+2, column =2)
        cell.value = value
        value = "0"
        cell = sheet.cell(row = i+2, column =3)
        cell.value = value

#Месяцы сдачи
mes=['Сентябрь','Октябрь','Ноябрь']
for i in range(len(namelist)):
    a=random.randint(0,2)
    if (a==0):
        value =mes[a]
        cell = sheet.cell(row = i+2, column =4)
        cell.value = value 
    if (a==1):
        value =mes[a]
        cell = sheet.cell(row = i+2, column =4)
        cell.value = value
    if (a==2):
        value =mes[a]
        cell = sheet.cell(row = i+2, column =4)
        cell.value = value

# добавляем второй листик
wb.create_sheet(title = 'Лагеря', index =1)

# заполняем второй листик
sheet = wb['Лагеря']
#Делаем строчечки больше ото слова не влезают,жирные они слишком
sheet.column_dimensions['A'].width = 25
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['C'].width = 15

value = "ФИО"
cell = sheet.cell(row = 1, column = 1)
cell.value = value

FIO ()

value = "Статус"
cell = sheet.cell(row = 1, column = 2)
cell.value = value

value = "Лагерь"
cell = sheet.cell(row = 1, column = 3)
cell.value = value

#Короче заполняем статусы и лагеря(как этоработает-я сама чуть не запуталась)
for i in range(len(namelist)):
    sheet = wb['Взносы']
    cell = sheet.cell(row = i+2, column =2)
    if (cell.value == "+"):
        sheet = wb['Лагеря']
        value = "Одобрено"
        cell = sheet.cell(row = i+2, column =2)
        cell.value = value
        sheet = wb['Взносы']
        cell = sheet.cell(row = i+2, column =4)
        if (cell.value == "Сентябрь"):
            sheet = wb['Лагеря']
            value = "Хогвартс"
            cell = sheet.cell(row = i+2, column =3)
            cell.value = value
        if (cell.value == "Октябрь"): 
            sheet = wb['Лагеря']
            value = "Алфея"
            cell = sheet.cell(row = i+2, column =3)
            cell.value = value
        if (cell.value == "Ноябрь"): 
            sheet = wb['Лагеря']
            value = "Нарния"
            cell = sheet.cell(row = i+2, column =3)
            cell.value = value
    else: 
        sheet = wb['Лагеря']
        value = "Не одобрено"
        cell = sheet.cell(row = i+2, column =2)
        cell.value = value
        value = "Дома тухнуть"
        cell = sheet.cell(row = i+2, column =3)
        cell.value = value
        
#Сохраняем сие творение
wb.save('ЯСДЕЛАЛЬ.xlsx')
